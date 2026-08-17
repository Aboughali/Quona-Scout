"""
Quona Case A -- Africa Fintech Sourcing System
ETL: reads the gold sheet and produces src/data/companies.json, funds.json
implementing the 5-stage funnel from specs/Quona_CaseA_Funnel_Filtration_Logic.docx.

Re-run any time the gold sheet is refreshed:
    python3 scripts/etl.py

Requires: openpyxl (pip install --break-system-packages openpyxl)
"""
import json
import re
from collections import defaultdict
from datetime import date, datetime

import openpyxl

TODAY = date(2026, 8, 12)
WINDOW_START = date(2023, 8, 1)

GOLD_SHEET = "data/Deals in africa Gold sheet - AMENDED JA V2.xlsx"

SERIES_A_PLUS = {
    "Series A", "Series A Extension", "Series B", "Pre-Series B", "Series B2",
    "Series C", "Series C Extension", "Series D", "Series E",
}
STILL_SEED_TYPES = {
    "Seed", "Pre-Seed", "Pre-seed", "Pre-Series A", "Seed Extension",
    "Bridge Round", "Venture Round", "Grant", "Debt", "M&A", "Private placement",
    "Green Bonds", "Secondary", "n.a",
}

GEO_AUTO_PASS_COUNTRIES = {"Egypt", "South Africa"}

FINTECH_SECTOR = "Fintech"
FINTECH_ADJACENT_KEYWORDS = [
    "payment", "lending", "credit", "embedded finance", "trade finance",
    "remittance", "wallet", "banking", "insurance", "insurtech", "loan",
    "financial infrastructure", "spend management", "expense management",
    "bnpl", "buy now pay later", "cross-border", "cards", "kyc",
    "open banking", "underwrit",
]

MULTI_MARKET_HINT_WORDS = [
    "pan-african", "pan african", "across africa", "multiple african markets",
    "expanding to", "operates in", "markets across", "several african countries",
]


_seen_slugs: dict = {}


def slugify(name: str) -> str:
    """Globally-unique slug for a CANONICAL company (called exactly once per canonical
    identity -- see norm_company_name-based grouping in main()). The numeric-suffix
    disambiguation here is now only a defensive fallback for the rare case where two
    genuinely DIFFERENT companies' names collapse to the same slug after stripping
    punctuation (e.g. "Bit.ly" vs "BitLy") -- it is no longer how case-variant duplicate
    rows of the SAME company (e.g. "CloudFret" / "Cloudfret") are told apart, since those
    are now merged into one company record before this is ever called."""
    base = re.sub(r"[^a-z0-9]+", "-", name.strip().lower()).strip("-")
    n = _seen_slugs.get(base, 0) + 1
    _seen_slugs[base] = n
    return base if n == 1 else f"{base}-{n}"


def load_rows(path, sheet_name):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, r)))
    return rows


def parse_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def split_investors(raw):
    if not raw:
        return []
    return [i.strip() for i in str(raw).split(",") if i.strip()]


INTAKE_SOURCE = "2023-2026 Intake Dataset"
HISTORY_SOURCE = "Deals 2019-2025"
BOTH_SOURCE = "Both (2023-2026 Intake Dataset + Deals 2019-2025)"


def norm_company_name(name: str) -> str:
    """Normalizes for cross-dataset matching: case, punctuation, and whitespace only --
    never treated as a new company just because the two sheets format it differently."""
    s = re.sub(r"[^a-z0-9]+", " ", str(name).strip().lower())
    return re.sub(r"\s+", " ", s).strip()


def as_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def round_fingerprint(date_val, rtype):
    """Identity key used to decide whether a round from the history sheet is the SAME
    round already present from the intake sheet (merge) or a genuinely separate one.
    Deliberately exact-date, not just year: some companies raise the same round *type*
    more than once in a year (e.g. two separate Debt tranches), and year-only matching
    would silently fold those two real, distinct raises into one -- exactly the kind of
    data loss this pipeline must not cause. A false negative (two near-identical rounds
    left unmerged because dates differ by a source's data-entry error) is far cheaper
    than a false positive that hides a real capital raise."""
    return (date_val.isoformat() if date_val else None, (rtype or "").strip().lower())


def build_round(r, source_dataset):
    return {
        "date": str(parse_date(r.get("Deal Date")) or ""),
        "type": r.get("Type"),
        "amountBracket": r.get("Bracket"),
        # Some cells hold literal text like "n.a" instead of being blank -- as_float()
        # normalizes that to a real null rather than letting a string leak into a field
        # the frontend treats as numeric (it crashed RoundsSection's total via string
        # concatenation: 0 + "n.a" -> "0n.a", which has no .toFixed).
        "amountUsdM": as_float(r.get("Amount raised $M")),
        "amountDisclosure": r.get("Amount disclosure"),
        "valuationUsdM": as_float(r.get("Valuation $M")),
        "estimatedValuationUsdM": as_float(r.get("Estimated Valuation (assuming 20% dilution)")),
        "investors": split_investors(r.get("Investors")),
        "link": r.get("Link to news"),
        "comment": r.get("Comment"),
        "sourceDataset": source_dataset,
        "conflict": None,
    }


def merge_historical_rounds(round_list, matched_history_rows, stats):
    """Enriches round_list (already built from the intake/window sheet) with rounds found
    for the same company in the Deals 2019-2025 sheet. Same-fingerprint rounds are merged
    (sources combined, investors unioned, amount/type disagreements flagged as a conflict --
    never silently overwritten) rather than duplicated; genuinely new historical rounds are
    appended. Mutates round_list in place and increments the shared `stats` counters dict."""
    fingerprints = {round_fingerprint(parse_date_str(rl["date"]), rl["type"]): rl for rl in round_list}

    for hr in matched_history_rows:
        h_date = parse_date(hr.get("Deal Date"))
        h_type = hr.get("Type")
        fp = round_fingerprint(h_date, h_type)
        h_investors = split_investors(hr.get("Investors"))
        h_amount = hr.get("Amount raised $M")

        existing = fingerprints.get(fp)
        if existing is not None:
            # Same round, seen in both datasets -- merge, don't duplicate.
            existing["sourceDataset"] = BOTH_SOURCE
            existing["investors"] = sorted(set(existing["investors"]) | set(h_investors))
            existing_amount = as_float(existing.get("amountUsdM"))
            h_amount_f = as_float(h_amount)
            if existing_amount is not None and h_amount_f is not None and abs(existing_amount - h_amount_f) > 0.01:
                existing["conflict"] = {
                    "field": "amountUsdM",
                    "values": [
                        {"value": existing_amount, "source": INTAKE_SOURCE},
                        {"value": h_amount, "source": HISTORY_SOURCE},
                    ],
                }
                stats["conflicts"] += 1
            stats["duplicates_merged"] += 1
        else:
            new_round = build_round(hr, HISTORY_SOURCE)
            round_list.append(new_round)
            fingerprints[fp] = new_round
            stats["historical_rounds_imported"] += 1

    round_list.sort(key=lambda r: r["date"] or "")


def dedupe_rounds_within_list(round_list, stats):
    """Folds exact-fingerprint (same date + same type) duplicate rounds together within a
    single list, unioning investors and flagging a genuine amount disagreement rather than
    picking one arbitrarily. This is what makes company-identity grouping safe: once two
    case-variant window-sheet rows for the same real company (e.g. "CloudFret" row +
    "Cloudfret" row) are combined into one company's round list, this collapses any round
    that both rows happened to report identically, while leaving rounds that only share a
    date (different type or amount) untouched -- per spec, same-stage-different-year (or
    same-date-different-type) rounds are kept separate unless the fingerprint is identical."""
    merged: list = []
    by_fingerprint: dict = {}
    for r in round_list:
        fp = round_fingerprint(parse_date_str(r["date"]), r["type"])
        existing = by_fingerprint.get(fp)
        if existing is not None:
            existing["investors"] = sorted(set(existing["investors"]) | set(r["investors"]))
            if existing["sourceDataset"] != r["sourceDataset"]:
                existing["sourceDataset"] = BOTH_SOURCE
            existing_amount = as_float(existing.get("amountUsdM"))
            r_amount = as_float(r.get("amountUsdM"))
            if existing_amount is not None and r_amount is not None and abs(existing_amount - r_amount) > 0.01:
                existing["conflict"] = {
                    "field": "amountUsdM",
                    "values": [
                        {"value": existing_amount, "source": existing.get("sourceDataset")},
                        {"value": r_amount, "source": r.get("sourceDataset")},
                    ],
                }
                stats["conflicts"] += 1
            stats["duplicates_merged"] += 1
        else:
            by_fingerprint[fp] = r
            merged.append(r)
    round_list[:] = merged


GENERIC_ROUND_TYPES = {"venture round", "n.a", ""}


def merge_type_conflicts(round_list, stats):
    """A second, narrower merge pass: two rounds on the exact same date with the exact
    same disclosed amount are almost certainly the same underlying event, just labeled
    with a different round *type* by each tracker (e.g. one sheet logs a raise as
    'Debt', the other as a generic 'Venture Round'). Without this, that shows up as a
    phantom second capital raise rather than one event with a type disagreement.

    The Deals 2019-2025 sheet systematically logs many rows as the generic 'Venture
    Round' where the intake sheet has since researched a specific stage (Pre-Seed,
    Series A, ...) -- that's reduced precision on one side, not a genuine factual
    disagreement, so it's resolved silently in favor of the specific label rather than
    flagged as a conflict. A conflict is only raised when BOTH sides name a specific,
    different type."""
    seen = {}
    merged_indices = set()
    for i, r in enumerate(round_list):
        amount = as_float(r.get("amountUsdM"))
        if amount is None or not r.get("date"):
            continue
        key = (r["date"], amount)
        if key in seen:
            j = seen[key]
            target = round_list[j]
            target_type = (target.get("type") or "").strip().lower()
            r_type = (r.get("type") or "").strip().lower()
            if target_type != r_type:
                target_generic = target_type in GENERIC_ROUND_TYPES
                r_generic = r_type in GENERIC_ROUND_TYPES
                if target_generic and not r_generic:
                    target["type"] = r.get("type")  # prefer the specific label
                elif not target_generic and not r_generic:
                    target["conflict"] = {
                        "field": "type",
                        "values": [
                            {"value": target.get("type"), "source": target.get("sourceDataset")},
                            {"value": r.get("type"), "source": r.get("sourceDataset")},
                        ],
                    }
                    stats["conflicts"] += 1
                # else: target already specific and r is generic -- keep target's type as-is
            target["investors"] = sorted(set(target["investors"]) | set(r["investors"]))
            if target["sourceDataset"] != r["sourceDataset"]:
                target["sourceDataset"] = BOTH_SOURCE
            merged_indices.add(i)
            stats["duplicates_merged"] += 1
        else:
            seen[key] = i
    if merged_indices:
        round_list[:] = [r for idx, r in enumerate(round_list) if idx not in merged_indices]


def parse_date_str(s):
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def is_fintech_adjacent(description: str) -> bool:
    if not description:
        return False
    d = description.lower()
    return any(kw in d for kw in FINTECH_ADJACENT_KEYWORDS)


def has_multi_market_hint(description: str) -> bool:
    if not description:
        return False
    d = description.lower()
    return any(kw in d for kw in MULTI_MARKET_HINT_WORDS)


def main():
    window_rows = load_rows(GOLD_SHEET, "Deals Aug 23-Aug26")
    history_rows = load_rows(GOLD_SHEET, "Deals 2019-2025")

    # Companies keyed by CANONICAL identity (normalized name), never the raw cell text --
    # this is the actual fix for duplicate company records. The gold sheet has a number of
    # case-variant rows for the same real company (e.g. "CloudFret" / "Cloudfret" rows,
    # each capturing a different one of that company's real rounds); grouping by raw
    # `name.strip()` treated each variant as its own company, which is the root cause audited
    # in Part 1 of the Quona Scout data-architecture review. One canonical key -> one company
    # record -> N round records, always.
    companies = defaultdict(list)
    for r in window_rows:
        name = r.get("Start-up name")
        if not name:
            continue
        companies[norm_company_name(name)].append(r)

    # Build a lookup: canonical name -> set of round types ever seen in EITHER sheet,
    # so Stage 3 "no announced Series A" checks full history, not just the window.
    all_round_types_by_name = defaultdict(set)
    for r in window_rows + history_rows:
        name = r.get("Start-up name")
        if not name:
            continue
        rtype = r.get("Type")
        if rtype:
            all_round_types_by_name[norm_company_name(name)].add(rtype)

    # Deals 2019-2025 rows grouped by normalized company name, for cross-dataset matching.
    # The 2023-2026 intake dataset alone determines the sourcing universe (Stage 1) -- this
    # is only ever used to ENRICH a company already in `companies`, never to add a new one.
    history_by_norm = defaultdict(list)
    for r in history_rows:
        name = r.get("Start-up name")
        if not name:
            continue
        history_by_norm[norm_company_name(name)].append(r)

    out_companies = []
    all_investor_names = set()
    recon = {
        "intake_companies": len(companies),
        "matched_to_history": 0,
        "not_matched": 0,
        "historical_rounds_imported": 0,
        "duplicates_merged": 0,
        "conflicts": 0,
        "name_normalizations_needed": 0,
    }

    for canonical_key, rows in companies.items():
        rounds_sorted = sorted(
            rows, key=lambda r: parse_date(r.get("Deal Date")) or WINDOW_START
        )
        first = rounds_sorted[0]
        # Canonical display name = earliest row's exact text -- deterministic, and consistent
        # with every other piece of company-level metadata already being sourced from `first`.
        name = (first.get("Start-up name") or "").strip()
        country = first.get("Country")
        region = first.get("Region")
        website = first.get("Website")
        description = first.get("Description") or ""
        sector = first.get("Sector")

        founding_year = first.get("Launch")
        founders = [
            first.get(k) for k in ("Founder 1 (CEO)", "Founder 2", "Founder 3", "Founder 4")
            if first.get(k)
        ]

        investors = set()
        round_list = []
        for r in rounds_sorted:
            rinvestors = split_investors(r.get("Investors"))
            investors.update(rinvestors)
            all_investor_names.update(rinvestors)
            round_list.append(build_round(r, INTAKE_SOURCE))

        # Multiple case-variant rows for this same canonical company can each report the
        # exact same round -- fold those together before enriching with history, so a
        # duplicate isn't mistaken for two separate financing events.
        dedupe_rounds_within_list(round_list, recon)

        # ---------- Enrich with Deals 2019-2025 historical rounds (matched companies only) ----------
        matched_history_rows = history_by_norm.get(canonical_key, [])
        if matched_history_rows:
            recon["matched_to_history"] += 1
            if any(hr.get("Start-up name", "").strip() != name for hr in matched_history_rows):
                recon["name_normalizations_needed"] += 1
            merge_historical_rounds(round_list, matched_history_rows, recon)
            merge_type_conflicts(round_list, recon)
            for hr in matched_history_rows:
                investors.update(split_investors(hr.get("Investors")))
                all_investor_names.update(split_investors(hr.get("Investors")))
        else:
            recon["not_matched"] += 1

        # ---------- Stage 1: All Raised ----------
        stage1_pass = True

        # ---------- Stage 2: Fintech Filter ----------
        is_core_fintech = sector == FINTECH_SECTOR
        adjacent_hit = (not is_core_fintech) and is_fintech_adjacent(description)
        stage2_pass = is_core_fintech
        stage2_reason = (
            f"Sector = {sector}" if is_core_fintech else
            (f"Sector = {sector}, but description matches fintech-adjacent keywords -- flagged for manual review, not auto-passed"
             if adjacent_hit else f"Sector = {sector}, no fintech signal")
        )

        # ---------- Stage 3: Stage Check (no Series A yet) ----------
        round_types = all_round_types_by_name.get(canonical_key, set())
        has_series_a_plus = bool(round_types & SERIES_A_PLUS)
        was_acquired = "M&A" in round_types
        stage3_pass = (not has_series_a_plus) and (not was_acquired)
        if has_series_a_plus:
            stage3_reason = f"Has raised: {', '.join(sorted(round_types & SERIES_A_PLUS))}"
        elif was_acquired:
            # Hard exclusion by analogy to the spec's shutdown rule: a company already
            # acquired is no longer an independent investable target, regardless of round status.
            stage3_reason = "Acquired (M&A event on record) -- no longer an independent investable target"
        else:
            stage3_reason = "No Series A or later round found in gold sheet history"

        # ---------- Stage 4: Geography Check ----------
        geo_auto_pass = country in GEO_AUTO_PASS_COUNTRIES
        multi_market_signal = has_multi_market_hint(description)
        stage4_pass = geo_auto_pass  # override applied later in the app, not baked in here
        stage4_reason = (
            f"HQ/base = {country} (auto-pass)" if geo_auto_pass else
            (f"HQ/base = {country}; description hints at multi-market operations -- candidate for manual 3+ markets override"
             if multi_market_signal else f"HQ/base = {country}; no multi-market evidence found")
        )

        # ---------- Passed all content gates (fintech + stage + geography) ----------
        # Formerly surfaced as "Stage 5 - Output Checkpoint"; that stage was removed because it
        # selected exactly the same set as Stage 4. Stage 5 is now the Syndicate Check, which is
        # computed live in the app (lib/resolve.ts), not here.
        passed_content_gates = stage2_pass and stage3_pass and stage4_pass

        # source confidence (rough proxy from disclosure + number of rounds with links)
        any_undisclosed = any(r.get("amountDisclosure") is None for r in round_list)
        has_link = any(r.get("link") for r in round_list)
        if has_link and not any_undisclosed:
            source_confidence = "High"
        elif has_link:
            source_confidence = "Medium"
        else:
            source_confidence = "Low"

        cut_stage = None
        cut_reason = None
        if not stage2_pass:
            cut_stage = 2
            cut_reason = stage2_reason
        elif not stage3_pass:
            cut_stage = 3
            cut_reason = stage3_reason
        elif not stage4_pass:
            cut_stage = 4
            cut_reason = stage4_reason

        out_companies.append({
            "id": slugify(name),
            "name": name,
            "website": website,
            "country": country,
            "region": region,
            "description": description,
            "sector": sector,
            "foundingYear": founding_year,
            "founders": founders,
            "investors": sorted(investors),
            "rounds": round_list,
            "latestRoundType": round_list[-1]["type"] if round_list else None,
            "sourceConfidence": source_confidence,
            "stage1": {"pass": stage1_pass, "reason": "Raised a round within Aug 2023-Aug 2026 window"},
            "stage2": {"pass": stage2_pass, "reason": stage2_reason, "adjacentCandidate": adjacent_hit},
            "stage3": {"pass": stage3_pass, "reason": stage3_reason},
            "stage4": {"pass": stage4_pass, "reason": stage4_reason, "autoPass": geo_auto_pass, "multiMarketSignal": multi_market_signal},
            "stage5": {"pass": None, "reason": "Pending syndicate research"},
            "overrides": {
                "seriesAReclass": {"active": False, "note": ""},
                "geography3Markets": {"active": False, "note": ""},
                "fintechPivot": {"active": False, "note": ""},
                "syndicateJudgment": {"active": False, "note": ""},
            },
            "cutStage": cut_stage,
            "cutReason": cut_reason,
        })

    out_companies.sort(key=lambda c: c["name"])

    with open("src/data/companies.json", "w") as f:
        json.dump(out_companies, f, indent=2)

    with open("src/data/investors_seen.json", "w") as f:
        json.dump(sorted(all_investor_names), f, indent=2)

    # Summary
    n1 = len(out_companies)
    n2 = sum(1 for c in out_companies if c["stage2"]["pass"])
    n3 = sum(1 for c in out_companies if c["stage2"]["pass"] and c["stage3"]["pass"])
    n4 = sum(1 for c in out_companies if c["stage2"]["pass"] and c["stage3"]["pass"] and c["stage4"]["pass"])
    print(f"Stage 1 (All Raised):      {n1}")
    print(f"Stage 2 (Fintech):         {n2}")
    print(f"Stage 3 (+ no Series A):   {n3}")
    print(f"Stage 4 (+ geography):     {n4}  <- pending Stage 5 syndicate gate")
    print(f"Unique investors touching Stage 4 candidates: "
          f"{len({inv for c in out_companies if c['stage2']['pass'] and c['stage3']['pass'] and c['stage4']['pass'] for inv in c['investors']})}")

    # Final surviving conflict count (a round can be merged more than once during the two
    # passes; only the LAST conflict state is kept, so re-count from final output rather
    # than trust the running increment, which counts merge *events* not surviving flags).
    recon["conflicts"] = sum(1 for c in out_companies for r in c["rounds"] if r.get("conflict"))

    print()
    print("=== Historical data reconciliation (Deals 2019-2025 -> 2023-2026 intake) ===")
    print(f"2023-2026 intake companies:              {recon['intake_companies']}")
    print(f"Matched to historical dataset:           {recon['matched_to_history']}")
    print(f"Not found in historical dataset:         {recon['not_matched']}")
    print(f"Historical rounds imported:              {recon['historical_rounds_imported']}")
    print(f"Duplicate rounds merged:                 {recon['duplicates_merged']}")
    print(f"Company-name matches needing normalization: {recon['name_normalizations_needed']}")
    print(f"Conflicting round information flagged:   {recon['conflicts']}")

    with open("src/data/reconciliation_report.json", "w") as f:
        json.dump(recon, f, indent=2)


if __name__ == "__main__":
    main()
